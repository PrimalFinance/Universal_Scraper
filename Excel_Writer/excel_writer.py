# OS level operations
import os

# For number storage and manipulation
import pandas as pd

# Excel writer utilities
import Excel_Writer.excel_writer_utils

# For excel access
import openpyxl as pyxl



class ExcelWriter:
    def __init__(self,ticker: str):
        self.ticker = ticker.upper()
        self.excel_file_name = f"{self.ticker}.xlsx"
        
        self.file_path = f"C:\\Users\William\\OneDrive\\Company Models CloudSave\\{ticker.upper()}"
        self.full_file_path = f"C:\\Users\\William\\OneDrive\\Company Models CloudSave\\{ticker.upper()}\\{self.excel_file_name}"

    '-------------------------------------------------------'
    def create_new_excel_file(self):
        # If the file does not exist.
        self.workbook = pyxl.Workbook(self.excel_file_name)
        self.workbook.create_sheet("Summary")
        self.workbook.create_sheet("Income Statement")
        self.workbook.create_sheet("Balance Sheet")
        self.workbook.create_sheet("Cash Flow")
        self.workbook.save(self.file_path)
        print(f"\n[Excel File Created] - {self.excel_file_name} was created with these sheets: {self.workbook.sheetnames}\n")

    '-------------------------------------------------------'
    def write_to_file(self, summary_df: pd.DataFrame, income_statement_df: pd.DataFrame, balance_sheet_df: pd.DataFrame, cash_flow_df: pd.DataFrame):

        print(f"Path: {self.file_path}")
        with pd.ExcelWriter(self.full_file_path) as writer:
            #main_df.to_excel(writer, "Main")
            #info_df.to_excel(writer, "Info")
            summary_df.to_excel(writer, "Summary")
            income_statement_df.to_excel(writer, "Income Statement")
            balance_sheet_df.to_excel(writer, "Balance Sheet")
            cash_flow_df.to_excel(writer, "Cash Flow")


    def test(self, summary_df):
        print(f"TEST: ")
        with pd.ExcelWriter(self.file_path) as writer:
            summary_df.to_excel(writer, "Summary")


    '-------------------------------------------------------'
    def check_if_file_exists(self, _path=None):
        if _path != None:
            return os.path.exists(_path)
        else:
            return os.path.exists(self.full_file_path)

    '-------------------------------------------------------'

    def get_worksheet(self, sheet_name: str):
        wb = pyxl.load_workbook(self.full_file_path)
        worksheet = wb[sheet_name]
        return worksheet

    '-------------------------------------------------------'

    def get_sheet_names(self):
        wb = pyxl.load_workbook(self.full_file_path)
        return wb.sheetnames

    '-------------------------------------------------------'
    '-------------------------------------------------------'
    '-------------------------------------------------------'
    '-------------------------------------------------------'

def main():

    s = ExcelWriter("AAPL")

   # s.get_worksheet("Balance Sheet")


