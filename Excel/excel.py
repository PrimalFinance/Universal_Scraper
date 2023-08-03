import pandas as pd
import openpyxl as pyxl


class CompanyModels:
    def __init__(self, ticker: str) -> None:

        self.excel_file_path = f"C:\\Users\\William\\OneDrive\\Company Models CloudSave\\{ticker}\\{ticker}.xlsx"

        print(f"File: {self.excel_file_path}")
        self.excel_file_name = self.excel_file_path.split("\\")[-1]

        print(f"File Path: {self.excel_file_path}")
        print(f"File Name: {self.excel_file_name}")
        try:
            self.excel_file = pyxl.load_workbook(self.excel_file_path)
        except KeyError:
            print(f"- [Error] {ticker}")
        '''-----------------------------------'''
    # This function will load the desired excel file and check if the sheet name in the parameter is present.
    # IF the sheet is not present, one will be created.
    # IF the sheet is present, nothing will happen.

    def create_sheet_if_not_exist(self, sheet_name: str):

        if sheet_name not in self.excel_file.sheetnames:
            self.excel_file.create_sheet(title=sheet_name)
            print(
                f"- [Created] '{sheet_name}' was created in {self.excel_file_name}.")
        else:
            print(
                f"- [Exists] '{sheet_name}' already in {self.excel_file_name}.")
        self.excel_file.save(self.excel_file_path)

    '''----------------------------------- Utilities -----------------------------------'''

    def organize_sheets(self, sheet_name: str, dest_index: int = 0):
        sheet_names = self.excel_file.sheetnames

        # Desired sheet order
        sheet_order = ["Main", "Info", "Model", "Summary",
                       "Income Statement", "Balance Sheet", "Cash Flow"]

        try:
            cur_sheet_index = sheet_names.index(sheet_name)
        except ValueError:
            self.create_sheet_if_not_exist(sheet_name)
            cur_sheet_index = self.excel_file.sheetnames.index(sheet_name)

        sheet_offset = self._calculate_offset(cur_sheet_index, dest_index)
        sheet = self.excel_file[sheet_name]
        self.excel_file.move_sheet(sheet, offset=sheet_offset)

        self.excel_file.save(self.excel_file_path)

    def _check_if_sheet_empty(self, sheet_name: str) -> bool:
        sheet = self.excel_file[sheet_name]

        df = pd.read_excel(self.excel_file_path, sheet_name=sheet_name)

        if df.empty:
            return True
        else:
            return False

    '''-----------------------------------'''

    def _calculate_offset(self, cur_index, dest_index) -> int:
        offset = cur_index - dest_index

        # If the current index is larger than the destination index, this means that we want to move it to the left.
        if cur_index > dest_index:
            offset *= -1

        return offset

    '''-----------------------------------'''
    # This function will data a sheetname, and write the data to the sheet, if it is empty.

    def write_df_to_sheet(self, sheet_name: str, data: pd.DataFrame, if_exists: str = "overlay"):
        # Check if the excel file is empty.
        empty = self._check_if_sheet_empty(sheet_name)
        if empty:
            with pd.ExcelWriter(self.excel_file_path, engine="openpyxl", mode="a", if_sheet_exists=if_exists) as writer:
                data.to_excel(writer, sheet_name=sheet_name, index=False)

    '''-----------------------------------'''

    def remove_empty_rows(self, sheet_name):
        workbook = pyxl.load_workbook(self.excel_file_path)

        worksheet = workbook[sheet_name]

        rows_to_delete = []

        for row in worksheet.iter_rows():
            if all(cell.value is None for cell in row):
                rows_to_delete.append(row[0].row)

        for row_index in reversed(rows_to_delete):
            worksheet.delete_rows(row_index)

        workbook.save(self.excel_file_path)

    '''-----------------------------------'''

    def if_sheet_item_exists(self, sheet_name: str, cell: str) -> bool:
        worksheet = self.excel_file[sheet_name]

        # If the value of the cell is blank.
        if worksheet[cell].value is None:
            return False
        # If there is data in the cell.
        else:
            return True

    def insert_sheet_item(self, sheet_name: str, cell: str, data) -> None:
        worksheet = self.excel_file[sheet_name]
        worksheet[cell] = data
        self.excel_file.save(self.excel_file_path)


'''----------------------------------------------------------------------------------------------------------------'''


class UniversalDashboard:
    def __init__(self, filepath=None) -> None:
        if filepath == None:
            self.excel_file_path = "C:\\Users\\William\\OneDrive\\Investing\\Universal_Dashboard.xlsx"
        else:
            self.excel_file_path = filepath
        self.excel_file_name = self.excel_file_path.split("\\")[-1]
        self.excel_file = pyxl.load_workbook(self.excel_file_path)

        self.col_dict = {'A': 1,
                         'B': 2,
                         'C': 3,
                         'D': 4,
                         'E': 5,
                         'F': 6,
                         'G': 7,
                         'H': 8,
                         'I': 9,
                         'J': 10,
                         'K': 11,
                         'L': 12,
                         'M': 13,
                         'N': 14,
                         'O': 15,
                         'P': 16,
                         'Q': 17,
                         'R': 18,
                         'S': 19,
                         'T': 20,
                         'U': 21,
                         'V': 22,
                         'W': 23,
                         'X': 24,
                         'Y': 25,
                         'Z': 26}
        self.row_start = 3

        self.columns = {"Basic Materials": "C",
                        "Communication Services": "D",
                        "Consumer Cyclical": "E",
                        "Consumer Defensive": "F",
                        "Energy": "G",
                        "Financial": "H",
                        "Healthcare": "I",
                        "Industrials": "J",
                        "Information Technology": "K",
                        "Real Estate": "L",
                        "Utilities": "M"}

        self.tickers = None

    '''-----------------------------------'''

    '''----------------------------------- Utilities -----------------------------------'''

    '''-----------------------------------'''

    '''-----------------------------------'''

    def test(self, sheet_name):

        df = pd.read_excel(self.excel_file_path, sheet_name=sheet_name)

        df = df.dropna()

        with pd.ExcelWriter(self.excel_file_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            df.to_excel(writer, sheet_name=sheet_name,
                        startrow=0, header=False, index=False)

    '''-----------------------------------'''
    '''-----------------------------------'''
    '''-----------------------------------'''

    '''-----------------------------------'''

    def set_ticker_in_dashboard(self) -> list:
        tickers = {}

        start_col = 2
        # Get basic materials tickers
        tickers["Basic Materials"] = self.get_column(col=start_col)
        # Get communication services
        tickers["Communication Services"] = self.get_column(col=start_col+1)
        # Get consumer cyclical
        tickers["Consumer Cyclical"] = self.get_column(col=start_col+2)
        # Get consumer defensive
        tickers["Consumer Defensive"] = self.get_column(col=start_col+3)
        # Get energy
        tickers["Energy"] = self.get_column(col=start_col+4)
        # Get finacial
        tickers["Financial"] = self.get_column(col=start_col+5)
        # Get healthcare
        tickers["Healthcare"] = self.get_column(col=start_col+6)
        # Get industrials
        tickers["Industrials"] = self.get_column(col=start_col+7)
        # Get Information Technology
        tickers["Information Technology"] = self.get_column(col=start_col+8)
        # Get real estate
        tickers["Real Estate"] = self.get_column(col=start_col+9)
        # Get utilities
        tickers["Utilities"] = self.get_column(col=start_col+10)

        self.tickers = tickers

    '''-----------------------------------'''

    def get_ticker_in_dashboard(self) -> dict:
        if self.tickers == None:
            self.set_ticker_in_dashboard()
        return self.tickers

    '''-----------------------------------'''

    def get_column(self, col: int):
        running = True
        tickers = []
        row = self.row_start
        index = 0
        for row in self.sheet.rows:
            # Skip the first row
            if index == 0:
                pass
            else:
                cell = row[col].value

                if cell == None:
                    break
                # Skip the sector name. Only grab the tickers.
                elif index > 1:
                    tickers.append(cell)

            # Increment loop variable
            index += 1
        return tickers

    '''-----------------------------------'''

    def get_company_sector(self, ticker):
        pass
    '''-----------------------------------'''

    def insert_tickers_to_dashboard(self, sector: str, data: list):
        worksheet = self.excel_file["Main"]
        col = self.columns[sector]
        row = self.row_start

        for l in range(len(data)):
            cell = f"{col}" + f"{row}"

            print(f"CELL: {cell}")

            cell_value = worksheet.cell(
                row=row, column=self.col_dict[col]).value

            if cell_value == data[l]['Ticker']:
                pass
            elif cell_value != data[l]['Ticker']:

                worksheet[cell] = data[l]['Ticker']

            row += 1
        self.excel_file.save(self.excel_file_path)
    '''-----------------------------------'''

    '''-----------------------------------'''
    '''-----------------------------------'''
    '''-----------------------------------'''
    '''-----------------------------------'''
